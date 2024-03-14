#This code is a simple Flask application that uses SQLAlchemy for database operations
#and Marshmallow for serialization. Let me break it down for you:
# serialization:marshmaloow can transfer the python object data  in to the serialize like json or xml
# use:dump for serialize the data 

from flask import Flask,jsonify,request
from flask_sqlalchemy import SQLAlchemy 
#(SQLAlchemy: An SQL toolkit and Object-Relational Mapping (ORM) library for Python.)
from flask_marshmallow import Marshmallow
# marshmallow is an object serialize and deseriazlize library
from openpyxl import load_workbook
# load work_book is a function from openpyxl library that load data from xl files 
app = Flask(__name__)
app.config["SQLALCHEMY_DATABASE_URI"] = 'sqlite:///db.sqllite3'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS']= False

db = SQLAlchemy(app)
ma = Marshmallow(app)

@app.route('/')
def hello():
    return ' hello world'

# creating a data base model:
class Fparser(db.Model):
    id=db.Column(db.Integer,primary_key = True)
    name=db.Column(db.String(200),nullable=False)
    age=db.Column(db.Integer,nullable=False)
    city=db.Column(db.String(200),nullable=False)


# parserschma class is a Marshmallow schema defining how the data should be serialized.
class parserschma(ma.Schema):
    class Meta:
        fields=['id','name','age','city']
parser_schema=parserschma()
parser_schemas=parserschma(many=True)
# we use many = true to indicates that schema
# can handle multiple object list of objects or multiple data while serializing

@app.route('/add',methods=['POST'])
def add_data():
    if request.method=='POST':
        excel_data=request.files['Mydata']
        Mydata=load_workbook(excel_data)
        newdata=Mydata.active
        for i in newdata.iter_rows(min_row=2,values_only=True):
            # this line is used to iterate over rows of an  exel file starting from second row
            #  and extracting the values

            # check_data=Fparser.filter_by(name=i[0],age=i[1],city=i[2])
            # print(check_data)
            # if check_data:
                # continue
            data = Fparser(name=i[0],age=i[1],city=i[2])
            db.session.add(data)
        db.session.commit()
    return "message:Data retrieve"    
    
@app.route('/get',methods=['GET'])
def get_all_data():
    all_post=Fparser.query.all()
    result = parser_schemas.dump(all_post)
    return jsonify(result)

@app.route('/get/<int:id>',methods=['GET'])
def get_data(id):
    post=Fparser.query.filter_by(id=id).first()
    result = parser_schema.dump(post)
    return jsonify(result)
        
@app.route('/update/<int:id>',methods=['PUT'])
def update_data(id):
    post=Fparser.query.get(id)
    name=request.json['name']
    age=request.json['age']
    city=request.json['city']

    post.name=name
    post.age=age
    post.city=city
    db.session.commit()
    return parser_schema.jsonify(post)

@app.route('/delete/<int:id>',methods=['DELETE'])
def delete_data(id):
    post=Fparser.query.get(id)
    db.session.delete(post)
    db.session.commit()
    return parser_schema.jsonify(post)


if __name__ =='__main__':
    app.run(debug=True)